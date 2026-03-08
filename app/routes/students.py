"""
University Gown Management System - Student Management Routes
"""

from flask import Blueprint, render_template, request, flash, redirect, url_for, Response
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from datetime import datetime
import os
import csv
from openpyxl import load_workbook

from app import db
from app.models import Student, AuditLog, Department
from app.forms import StudentForm
from app.access_control import admin_required, superadmin_required

students_bp = Blueprint('students', __name__, url_prefix='/students')


@students_bp.route('/')
@login_required
@admin_required
def index():
    """List all students (Admin & SuperAdmin)"""
    page = request.args.get('page', 1, type=int)
    per_page = 20
    
    # Search functionality
    search = request.args.get('search', '')
    search_type = request.args.get('search_type', 'index')
    
    query = Student.query
    
    # Filter by department if user is not a superadmin
    if not current_user.is_superadmin() and current_user.department_id:
        query = query.filter(Student.department_id == current_user.department_id)
    
    if search:
        if search_type == 'index':
            query = query.filter(Student.index_number.ilike(f'%{search}%'))
        elif search_type == 'name':
            query = query.filter(Student.full_name.ilike(f'%{search}%'))
    
    students = query.order_by(Student.index_number).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    return render_template('students/index.html', students=students, search=search, search_type=search_type)


@students_bp.route('/new', methods=['GET', 'POST'])
@login_required
@superadmin_required
def new():
    """Create new student - SuperAdmin only"""
    form = StudentForm()
    
    if form.validate_on_submit():
        # Create student with only essential fields
        student = Student(
            index_number=form.index_number.data,
            full_name=form.full_name.data,
            programme=form.programme.data,
            level=form.level.data
        )
        
        # Auto-assign department based on programme
        student.assign_department()
        
        db.session.add(student)
        db.session.commit()
        
        # Log the action
        log = AuditLog(
            user_id=current_user.id,
            action='Create Student',
            entity_type='Student',
            entity_id=student.id,
            details=f'Created student {student.index_number}',
            ip_address=request.remote_addr
        )
        db.session.add(log)
        db.session.commit()
        
        flash(f'Student {student.full_name} created successfully!', 'success')
        return redirect(url_for('students.index'))
    
    return render_template('students/form.html', form=form, title='Add Student')


@students_bp.route('/<int:student_id>')
@login_required
def view(student_id):
    """View student details"""
    student = Student.query.get_or_404(student_id)
    
    # Get student's transaction history
    transactions = student.transactions.order_by(db.desc('created_at')).all()
    
    return render_template('students/view.html', student=student, transactions=transactions)


@students_bp.route('/<int:student_id>/edit', methods=['GET', 'POST'])
@login_required
@superadmin_required
def edit(student_id):
    """Edit student - SuperAdmin only"""
    student = Student.query.get_or_404(student_id)
    form = StudentForm(obj=student)
    
    # Custom validation for edit (ignore current index_number)
    def validate_index_number(self, index_number):
        existing = Student.query.filter(
            Student.index_number == index_number.data,
            Student.id != student.id
        ).first()
        if existing:
            from wtforms import ValidationError
            raise ValidationError('Index number already exists')
    
    if form.validate_on_submit():
        student.index_number = form.index_number.data
        student.full_name = form.full_name.data
        student.programme = form.programme.data
        student.level = form.level.data
        
        # Auto-assign department based on programme
        student.assign_department()
        
        db.session.commit()
        
        # Log the action
        log = AuditLog(
            user_id=current_user.id,
            action='Edit Student',
            entity_type='Student',
            entity_id=student.id,
            details=f'Edited student {student.index_number}',
            ip_address=request.remote_addr
        )
        db.session.add(log)
        db.session.commit()
        
        flash(f'Student {student.full_name} updated successfully!', 'success')
        return redirect(url_for('students.index'))
    
    return render_template('students/form.html', form=form, title=f'Edit Student: {student.index_number}')


@students_bp.route('/<int:student_id>/delete', methods=['POST'])
@login_required
@superadmin_required
def delete(student_id):
    """Delete student - SuperAdmin only"""
    student = Student.query.get_or_404(student_id)
    
    # Check if student has active rentals
    if student.has_active_rental():
        flash('Cannot delete student with active rentals.', 'danger')
        return redirect(url_for('students.index'))
    
    index_number = student.index_number
    
    # Log before deletion
    log = AuditLog(
        user_id=current_user.id,
        action='Delete Student',
        entity_type='Student',
        entity_id=student.id,
        details=f'Deleted student {index_number}',
        ip_address=request.remote_addr
    )
    db.session.add(log)
    
    db.session.delete(student)
    db.session.commit()
    
    flash(f'Student {index_number} deleted successfully!', 'success')
    return redirect(url_for('students.index'))


@students_bp.route('/import', methods=['GET', 'POST'])
@login_required
@superadmin_required
def import_students():
    """Import students from Excel/CSV - SuperAdmin only"""
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file selected', 'danger')
            return redirect(request.url)
        
        file = request.files['file']
        
        if file.filename == '':
            flash('No file selected', 'danger')
            return redirect(request.url)
        
        if file:
            filename = secure_filename(file.filename)
            filepath = os.path.join('/tmp', filename)
            file.save(filepath)
            
            try:
                # Read Excel or CSV
                rows_data = []
                if filename.endswith('.csv'):
                    with open(filepath, 'r', encoding='utf-8') as f:
                        reader = csv.DictReader(f)
                        rows_data = list(reader)
                else:
                    wb = load_workbook(filepath, read_only=True)
                    ws = wb.active
                    headers = [cell.value for cell in ws[1]]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        row_dict = {headers[i]: row[i] for i in range(len(headers))}
                        rows_data.append(row_dict)
                    wb.close()
                
                # Normalize column names - strip whitespace and convert to lowercase
                normalized_headers = {}
                for header in rows_data[0].keys():
                    if header is not None:  # Skip None headers
                        normalized_headers[header.strip().lower()] = header
                
                # Define required columns (support both space and underscore formats)
                required_columns = ['index_number', 'full_name', 'programme', 'level']
                # Also support "Index Number", "Full Name", "Programme", "Level" format
                alternative_names = {
                    'index number': 'index_number',
                    'full name': 'full_name', 
                    'programme': 'programme',
                    'level': 'level'
                }
                
                # Get available columns - handle None values
                available_columns = []
                for col in rows_data[0].keys():
                    if col is not None and col.strip():
                        available_columns.append(col.strip().lower())
                
                print(f"DEBUG: Available columns: {available_columns}")
                
                # Simple keyword-based matching - be very lenient
                def find_matching_column(keywords):
                    """Find any column that contains any of the keywords"""
                    for col in available_columns:
                        col_clean = col.replace(' ', '').replace('_', '').replace('-', '')
                        for kw in keywords:
                            kw_clean = kw.replace(' ', '').replace('_', '').replace('-', '')
                            if kw_clean in col_clean:
                                return col
                    return None
                
                # Map keywords to internal field names
                field_mapping = {
                    'index': 'index_number',
                    'name': 'full_name',
                    'program': 'programme',
                    'course': 'programme',
                    'major': 'programme',
                    'level': 'level',
                    'year': 'level'
                }
                
                # Check if we have at least some columns
                has_index = find_matching_column(['index', 'student', 'registration', 'id'])
                has_name = find_matching_column(['name', 'full'])
                has_program = find_matching_column(['program', 'course', 'major'])
                has_level = find_matching_column(['level', 'year', 'class'])
                
                print(f"DEBUG: has_index={has_index}, has_name={has_name}, has_program={has_program}, has_level={has_level}")
                
                # Only fail if完全没有匹配
                if not (has_index and has_name):
                    flash(f'Could not find required columns (Index/Name). Found: {", ".join(available_columns)}', 'danger')
                    return redirect(url_for('students.import_students'))
                
                # Import students
                imported = 0
                duplicates = 0
                errors = []
                
                for idx, row in enumerate(rows_data, start=2):
                    try:
                        # Normalize row keys
                        normalized_row = {}
                        for k, v in row.items():
                            if k is not None:
                                normalized_row[k.strip()] = v
                        
                        # Simple keyword-based value extraction
                        def find_value(keywords):
                            """Find value using keyword matching"""
                            for key in normalized_row.keys():
                                key_clean = key.lower().replace(' ', '').replace('_', '').replace('-', '')
                                for kw in keywords:
                                    kw_clean = kw.lower().replace(' ', '').replace('_', '').replace('-', '')
                                    if kw_clean in key_clean:
                                        return normalized_row[key]
                            return None
                        
                        # Get values - use more specific keywords
                        index_value = find_value(['index', 'student', 'registration', 'id'])
                        name_value = find_value(['name', 'fullname'])
                        
                        # For programme, match program/course/major
                        prog_value = find_value(['program', 'course', 'major'])
                        
                        # For level, check if there's a dedicated level column
                        level_value = None
                        for key in normalized_row.keys():
                            key_clean = key.lower().replace(' ', '').replace('_', '').replace('-', '')
                            # Only match if it's specifically a level column (not programme)
                            if any(kw in key_clean for kw in ['level', 'yr', 'year', 'classlevel']):
                                level_value = normalized_row[key]
                                break
                        
                        # If still no level, check if programme contains level info
                        if not level_value and prog_value:
                            prog_str = str(prog_value).lower()
                            if 'level 100' in prog_str or 'l100' in prog_str or 'year 1' in prog_str:
                                level_value = 'Level 100'
                            elif 'level 200' in prog_str or 'l200' in prog_str or 'year 2' in prog_str:
                                level_value = 'Level 200'
                            elif 'level 300' in prog_str or 'l300' in prog_str or 'year 3' in prog_str:
                                level_value = 'Level 300'
                            elif 'level 400' in prog_str or 'l400' in prog_str or 'year 4' in prog_str:
                                level_value = 'Level 400'
                            elif 'diploma' in prog_str:
                                level_value = 'Diploma'
                            elif 'master' in prog_str or 'msc' in prog_str or 'm.sc' in prog_str:
                                level_value = 'Master'
                            elif 'phd' in prog_str or 'doctorate' in prog_str:
                                level_value = 'PhD'
                            elif 'bachelor' in prog_str or 'bsc' in prog_str or 'b.sc' in prog_str or 'ba' in prog_str:
                                level_value = 'Bachelor'
                        
                        # Skip if index or name is missing
                        if not index_value or not name_value:
                            errors.append(f"Row {idx}: Missing index or name")
                            continue
                        
                        # Use defaults for optional fields
                        if not prog_value:
                            prog_value = "Unknown Programme"
                        if not level_value:
                            level_value = "Level 100"
                        
                        # Check for duplicates
                        existing = Student.query.filter_by(index_number=str(index_value).strip()).first()
                        
                        if existing:
                            duplicates += 1
                            continue
                        
                        # Create student with only essential fields
                        student = Student(
                            index_number=str(index_value).strip(),
                            full_name=str(name_value).strip(),
                            programme=str(prog_value).strip(),
                            level=str(level_value).strip()
                        )
                        
                        # Auto-assign department based on programme
                        student.assign_department()
                        
                        db.session.add(student)
                        imported += 1
                        
                    except Exception as e:
                        errors.append(f"Row {idx}: {str(e)}")
                
                db.session.commit()
                
                # Log the action
                log = AuditLog(
                    user_id=current_user.id,
                    action='Import Students',
                    entity_type='Student',
                    details=f'Imported {imported} students, {duplicates} duplicates skipped',
                    ip_address=request.remote_addr
                )
                db.session.add(log)
                db.session.commit()
                
                # Clean up
                os.remove(filepath)
                
                flash(f'Successfully imported {imported} students. {duplicates} duplicates skipped.', 'success')
                
                if errors:
                    flash(f'Errors: {", ".join(errors[:5])}', 'warning')
                
            except Exception as e:
                flash(f'Error processing file: {str(e)}', 'danger')
    
    return render_template('students/import.html')


@students_bp.route('/search')
@login_required
def search():
    """Quick student search (JSON)"""
    query = request.args.get('q', '')
    
    if len(query) < 2:
        return {'results': []}
    
    # Build base query
    base_query = Student.query
    
    # Filter by department if user is not a superadmin
    if not current_user.is_superadmin() and current_user.department_id:
        base_query = base_query.filter(Student.department_id == current_user.department_id)
    
    students = base_query.filter(
        (Student.index_number.ilike(f'%{query}%')) | 
        (Student.full_name.ilike(f'%{query}%'))
    ).limit(10).all()
    
    results = [{
        'id': s.id,
        'index_number': s.index_number,
        'full_name': s.full_name,
        'programme': s.programme,
        'level': s.level,
        'has_rental': s.has_active_rental()
    } for s in students]
    
    return {'results': results}


@students_bp.route('/lookup/<index_number>')
@login_required
def lookup(index_number):
    """Look up student by index number"""
    student = Student.query.filter_by(index_number=index_number).first()
    
    if not student:
        return {'error': 'Student not found'}, 404
    
    return {
        'id': student.id,
        'index_number': student.index_number,
        'full_name': student.full_name,
        'programme': student.programme,
        'level': student.level,
        'department': student.department.name if student.department else None,
        'has_active_rental': student.has_active_rental()
    }
